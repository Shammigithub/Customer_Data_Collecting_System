from tkinter import *
from datetime import date
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import pathlib
from time import strftime
from time import *

background = "#062831"
framebg = "#EDEDED"
framefg = "#062831"

root = Tk()
root.title("Customer Information System")
root.geometry("1250x750+0+00")
root.iconbitmap('customer.ico')
root.config(bg=background)

file = pathlib.Path('Customer_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Gender"
    sheet['D1'] = "NIC"
    sheet['E1'] = "Date of Come"
    sheet['F1'] = "Time"
    sheet['G1'] = "Address"
    sheet['H1'] = "Tel. No."

    file.save('Customer_data.xlsx')


# Exit
def Exit():
    root.destroy()


# Sign in
def ok():
    def login():
        email = entry1.get()
        password = entry2.get()

        if (email == "" and password == ""):
            messagebox.showinfo("", "Blank not allowed")

        elif (email == "asankadilshan44098@gmail.com" and password == "12345"):  # asankadilshan44098@gmail.com
            messagebox.showinfo("", "Login success")
            root2.withdraw()

            root3 = Tk()
            root3.title("Customer Information System")
            root3.geometry("1250x750+0+00")
            root3.iconbitmap('customer.ico')
            root3.config(bg=background)

            # top frames
            Label(root3, text='', width=10, height=1, bg="#429683", font="arial 28 bold").pack(side=TOP, fill=X)
            Label(root3, text='AD Grossery Shop', width=10, height=1, bg="#429683", font="arial 30 bold").pack(side=TOP,
                                                                                                               fill=X)
            Label(root3, text='!..Welcome..!', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
                side=TOP, fill=X)

            def search():
                search_date = date_entry.get()

                # Clear the search results
                search_results.delete(*search_results.get_children())

                for row in sheet.iter_rows(values_only=True):
                    if row[4] == search_date:
                        search_results.insert("", "end", values=row)

            # Load the workbook
            workbook = load_workbook('Customer_data.xlsx')
            # Get the active sheet
            sheet = workbook.active

            date_entry = Entry(root3, width=30)
            date_entry.pack(pady=5)

            search_button = Button(root3, text="Search", bg="#93E0EA", font='arial 10 bold', width=20, height=2,
                                   command=search)
            search_button.pack(pady=5)

            search_results_frame = Frame(root3)
            search_results_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)

            search_results = ttk.Treeview(search_results_frame)
            search_results["columns"] = (
                "Registration No.", "Name", "Gender", "NIC", "Date of Come", "Time", "Address", "Tel. No.")

            # Configure columns
            search_results.column("#0", width=0, stretch=NO)
            search_results.column("Registration No.", width=100)
            search_results.column("Name", width=100)
            search_results.column("Gender", width=100)
            search_results.column("NIC", width=100)
            search_results.column("Date of Come", width=100)
            search_results.column("Time", width=100)
            search_results.column("Address", width=100)
            search_results.column("Tel. No.", width=100)

            # Create headings
            search_results.heading("#0", text="", anchor=W)
            search_results.heading("Registration No.", text="Registration No.")
            search_results.heading("Name", text="Name")
            search_results.heading("Gender", text="Gender")
            search_results.heading("NIC", text="NIC")
            search_results.heading("Date of Come", text="Date of Come")
            search_results.heading("Time", text="Time")
            search_results.heading("Address", text="Address")
            search_results.heading("Tel. No.", text="Tel. No.")

            search_results.pack(side=LEFT, fill=BOTH, expand=True)

            # Add scrollbar to the right of the treeview
            scrollbar = Scrollbar(search_results_frame, orient=VERTICAL, command=search_results.yview)
            scrollbar.pack(side=RIGHT, fill=Y)
            search_results.configure(yscrollcommand=scrollbar.set)

            exit_button = Button(root3, text="Exit", bg="#93CAEA", font='arial 10 bold', width=20, height=2,
                                 command=exit)
            exit_button.pack(pady=5)

            # bottom frames
            Label(root3, text='', width=10, height=1, bg="#429683", ).pack(side=BOTTOM, fill=X)
            Label(root3, text='077-0867965', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM,
                                                                                                          fill=X)
            Label(root3, text='Galle', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
            Label(root3, text='Karagoda', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM,
                                                                                                       fill=X)
            Label(root3, text='Mulanagoda,', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM,
                                                                                                          fill=X)

            Label(root3, text='Thank You Come Again', width=10, height=2, bg="#70BBAA", fg='#fff',
                  font='arial 20 bold').pack(
                side=BOTTOM, fill=X)



        else:
            messagebox.showinfo("", "Incorrect username and password")

    def Exit1():
        root2.destroy()

    root2 = Tk()
    root2.title("Customer Information System")
    root2.geometry("1250x750+0+00")
    root2.iconbitmap('customer.ico')
    root2.config(bg=background)

    global entry1
    global entry2

    # top frames
    Label(root2, text='', width=10, height=1, bg="#429683", font="arial 28 bold").pack(side=TOP, fill=X)
    Label(root2, text='AD Grossery Shop', width=10, height=1, bg="#429683", font="arial 30 bold").pack(side=TOP, fill=X)
    Label(root2, text='!..Welcome..!', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
        side=TOP, fill=X)

    # bottom frames
    Label(root2, text='', width=10, height=1, bg="#429683", ).pack(side=BOTTOM, fill=X)
    Label(root2, text='077-0867965', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
    Label(root2, text='Galle', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
    Label(root2, text='Karagoda', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
    Label(root2, text='Mulanagoda,', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)

    Label(root2, text='Thank You Come Again', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
        side=BOTTOM, fill=X)

    # Sign in Window
    obj = LabelFrame(root2, text="User Login", font=20, bd=2, width=345, bg=framebg, fg=framefg, height=300, )
    obj.place(x=460, y=220)
    Label(root2, text="E-mail", width=8).place(x=500, y=280)
    Label(root2, text="Password", width=8).place(x=500, y=330)

    Email = StringVar()
    entry1 = Entry(obj, textvariable=Email, width=30, font="arial 10")
    entry1.place(x=110, y=40)

    Password = StringVar()
    entry2 = Entry(obj, textvariable=Password, width=30, font="arial 10")
    entry2.place(x=110, y=90)
    entry2.config(show="*")

    Button(root2, text="Sign in", width=12, height=2, font="arial 12 bold", bg="#81CBEB", command=login).place(x=490,
                                                                                                               y=400)
    Button(root2, text="Exit", width=12, height=2, font="arial 12 bold", bg="#93CAEA", command=Exit1).place(x=645,
                                                                                                            y=400)
    root.withdraw()

    root2.mainloop()


def registration_no():
    file = openpyxl.load_workbook('Customer_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value
    print(max_row_value)

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set("1")


def Clear():
    Name.set('')
    NIC.set('')
    Time.set('')
    Address.set('')
    Tel.set('')


    registration_no()

    saveButton.config(state='normal')


# Save
def Save():
    Rn1 = Registration.get()
    N1 = Name.get()

    try:
        G1 = gender
    except:
        messagebox.showerror("error", "Select Gender!")

    N2 = NIC.get()
    D1 = Date.get()
    T1 = Time.get()
    A1 = Address.get()
    T2 = Tel.get()

    if N1 == "" or N2 == "" or T1 == "" or A1 == "" or T2 == "":
        messagebox.showerror("error", "Few Data is missing!")

    else:
        file = openpyxl.load_workbook('Customer_data.xlsx')
        sheet = file.active

        sheet.cell(column=1, row=sheet.max_row + 1, value=Rn1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=G1)
        sheet.cell(column=4, row=sheet.max_row, value=N2)
        sheet.cell(column=5, row=sheet.max_row, value=D1)
        sheet.cell(column=6, row=sheet.max_row, value=T1)
        sheet.cell(column=7, row=sheet.max_row, value=A1)
        sheet.cell(column=8, row=sheet.max_row, value=T2)

        file.save('Customer_data.xlsx')

        messagebox.showinfo("info", "Sucessfully data entered!!!")

        Clear()

        registration_no()


# gender
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"

    else:
        gender = "Female"


# top frames
Label(root, text='', width=10, height=1, bg="#429683", font="arial 28 bold").pack(side=TOP, fill=X)
Label(root, text='AD Grossery Shop', width=10, height=1, bg="#429683", font="arial 30 bold").pack(side=TOP, fill=X)
Label(root, text='!..Welcome..!', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
    side=TOP, fill=X)

# bottom frames
Label(root, text='', width=10, height=1, bg="#429683", ).pack(side=BOTTOM, fill=X)
Label(root, text='077-0867965', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
Label(root, text='Galle', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
Label(root, text='Karagoda', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
Label(root, text='Mulanagoda,', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)

Label(root, text='Thank You Come Again', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
    side=BOTTOM, fill=X)

# Registration and Date
Label(root, text="Registration No.", font='arial 13', fg=framebg, bg=background).place(x=30, y=200)
Label(root, text="Date", font='arial 13', fg=framebg, bg=background).place(x=500, y=200)
Label(root, text="Time", font='arial 13', fg=framebg, bg=background).place(x=750, y=200)

Registration = IntVar()
Date = StringVar()
Time = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font='arial 10')
reg_entry.place(x=160, y=200)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font='arial 10')
date_entry.place(x=550, y=200)

Date.set(d1)

time_label = Label(root, text="", font='arial 10', width=15)
time_label.place(x=800, y=200)


def update_time():
    current_time = strftime("%H:%M:%S %p")
    time_label.config(text=current_time)
    root.after(1000, update_time)


update_time()

# Customer details
obj = LabelFrame(root, text="Customer's Details", font=20, bd=3, width=900, bg=framebg, fg=framefg, height=250,
                 relief=GROOVE)
obj.place(x=50, y=250)

Label(root, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=90, y=320)
Label(root, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=90, y=370)
Label(root, text="NIC:", font="arial 13", bg=framebg, fg=framefg).place(x=90, y=420)

Label(root, text="Address:", font="arial 13", bg=framebg, fg=framefg).place(x=550, y=320)
Label(root, text="Time:", font="arial 13", bg=framebg, fg=framefg).place(x=550, y=370)
Label(root, text="Tel. No.:", font="arial 13", bg=framebg, fg=framefg).place(x=550, y=420)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=30, font="arial 10")
name_entry.place(x=180, y=50)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=180, y=100)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=240, y=100)

NIC = StringVar()
nic_entry = Entry(obj, textvariable=NIC, width=20, font="arial 10")
nic_entry.place(x=180, y=150)

time_entry = Entry(obj, textvariable=Time, width=20, font="arial 10")
time_entry.place(x=630, y=100)

Address = StringVar()
address_entry = Entry(obj, textvariable=Address, width=30, font="arial 10")
address_entry.place(x=630, y=50)

Tel = StringVar()
tel_entry = Entry(obj, textvariable=Tel, width=20, font="arial 10")
tel_entry.place(x=630, y=150)

# button
Button(root, text="Sign in", width=19, height=2, font="arial 12 bold", bg="#81CBEB", command=ok).place(x=1000, y=215)

saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="#9BBEEA", command=Save)
saveButton.place(x=1000, y=295)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="#93E0EA", command=Clear).place(x=1000, y=375)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="#93CAEA", command=Exit).place(x=1000, y=455)

root.mainloop()
