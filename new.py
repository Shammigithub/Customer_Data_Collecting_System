from tkinter import *
from datetime import date
from tkinter import filedialog, ttk
from tkinter import messagebox

import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
from time import strftime
import pandas as pd
from time import *

background = "#062831"
framebg = "#EDEDED"
framefg = "#062831"

root = Tk()
root.title("Customer Information System")
root.geometry("1250x700+0+00")
root.config(bg=background)

file = pathlib.Path('Customer_data.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Registration No."
    sheet['B1'] = "Name"
    sheet['C1'] = "Gender"
    sheet['D1'] = "NIC"
    sheet['E1'] = "Date of Come"
    sheet['F1'] = "Time"
    sheet['G1'] = "Address"
    sheet['H1'] = "Tel. No."

    file.save('Customer_data.xlsx')


# Exit
def Exit():
    root.destroy()


def ok():
    def login():
        email = entry1.get()
        password = entry2.get()

        if (email == "" and password == ""):
            messagebox.showinfo("", "Blank not allowed")

        elif (email == "asankadilshan44098@gmail.com" and password == "12345"):
            messagebox.showinfo("", "Login success")
            root2.withdraw()

            root3 = Tk()
            root3.title("Customer Information System")
            root3.geometry("1250x700+0+00")
            root3.config(bg=background)

            # def search():

            # top frames
            Label(root3, text='', width=10, height=1, bg="#429683", font="arial 28 bold").pack(side=TOP, fill=X)
            Label(root3, text='AD Store', width=10, height=1, bg="#429683", font="arial 30 bold").pack(side=TOP, fill=X)
            Label(root3, text='!..Welcome..!', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
                side=TOP, fill=X)

            # Search
            Search = StringVar()
            e1 = Entry(root3, textvariable=Search, width=15, bd=2, font="arial 20").place(x=820, y=70)
            Srch = Button(root3, text="Search", compound=LEFT, width=123, bg='#68ddfa',
                          font='arial 13 bold', )  # command=search)
            Srch.place(x=1060, y=66)



            # bottom frames
            Label(root3, text='', width=10, height=1, bg="#429683", ).pack(side=BOTTOM, fill=X)
            Label(root3, text='077-0867965', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM,
                                                                                                         fill=X)
            Label(root3, text='Galle', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
            Label(root3, text='Karagoda', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM,
                                                                                                      fill=X)
            Label(root3, text='Mulanagoda,', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM,
                                                                                                         fill=X)

            Label(root3, text='Thank You Come Again', width=10, height=2, bg="#70BBAA", fg='#fff',
                  font='arial 20 bold').pack(
                side=BOTTOM, fill=X)



        else:
            messagebox.showinfo("", "Incorrect username and password")

    def Exit1():
        root2.destroy()

    root2 = Tk()
    root2.title("Customer Information System")
    root2.geometry("1250x700+0+00")
    root2.config(bg=background)

    global entry1
    global entry2

    # top frames
    Label(root2, text='', width=10, height=1, bg="#429683", font="arial 28 bold").pack(side=TOP, fill=X)
    Label(root2, text='AD Store', width=10, height=1, bg="#429683", font="arial 30 bold").pack(side=TOP, fill=X)
    Label(root2, text='!..Welcome..!', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
        side=TOP, fill=X)

    # bottom frames
    Label(root2, text='', width=10, height=1, bg="#429683", ).pack(side=BOTTOM, fill=X)
    Label(root2, text='077-0867965', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
    Label(root2, text='Galle', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
    Label(root2, text='Karagoda', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)
    Label(root2, text='Mulanagoda,', width=10, height=1, bg="#429683", font='arial 10 bold').pack(side=BOTTOM, fill=X)

    Label(root2, text='Thank You Come Again', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
        side=BOTTOM, fill=X)

    # Sign in Window
    obj =  LabelFrame(root2, text="User Login", font=20, bd=2, width=345, bg=framebg, fg=framefg, height=300,)
    obj.place(x=460, y=200)
    Label(root2, text="E-mail", width=8).place(x=500, y=280)
    Label(root2, text="Password", width=8).place(x=500, y=330)

    Email = StringVar()
    entry1 = Entry(obj, textvariable=Email, width=30, font="arial 10")
    entry1.place(x=120, y=55)

    Password = StringVar()
    entry2 = Entry(obj, textvariable=Password, width=30, font="arial 10")
    entry2.place(x=120, y=105)
    entry2.config(show="*")

    Button(root2, text="Sign in", width=10, height=1, font="arial 12 bold", bg="#81CBEB", command=login).place(x=640,
                                                                                                                 y=395)
    Button(root2, text="Exit", width=10, height=1, font="arial 12 bold", bg="#93CAEA", command=Exit1).place(x=510,
                                                                                                              y=395)
    root.withdraw()

    root2.mainloop()


# time
# def my_time():
# time_string = strftime("%H:%M:%S %p")
# te_str.set(time_string)
# my_time = Time.my_time()
# t1 = my_time.strftime("%H:%M:%S %p")

# Time.set(t1)
# Registration No.
def registration_no():
    file = openpyxl.load_workbook('Customer_data.xlsx')
    sheet = file.active
    row = sheet.max_row

    max_row_value = sheet.cell(row=row, column=1).value
    print(max_row_value)

    try:
        Registration.set(max_row_value + 1)

    except:
        Registration.set("1")


def Clear():
    Name.set('')
    NIC.set('')
    Time.set('')
    Address.set('')
    Tel.set('')

    # Gender.set("Select Gender")

    registration_no()

    saveButton.config(state='normal')


# Save
def Save():
    Rn1 = Registration.get()
    N1 = Name.get()

    try:
        G1 = gender
    except:
        messagebox.showerror("error", "Select Gender!")

    N2 = NIC.get()
    D1 = Date.get()
    T1 = Time.get()
    A1 = Address.get()
    T2 = Tel.get()

    if N1 == "" or N2 == "" or T1 == "" or A1 == "" or T2 == "":
        messagebox.showerror("error", "Few Data is missing!")

    else:
        file = openpyxl.load_workbook('Customer_data.xlsx')
        sheet = file.active

        sheet.cell(column=1, row=sheet.max_row + 1, value=Rn1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=G1)
        sheet.cell(column=4, row=sheet.max_row, value=N2)
        sheet.cell(column=5, row=sheet.max_row, value=D1)
        sheet.cell(column=6, row=sheet.max_row, value=T1)
        sheet.cell(column=7, row=sheet.max_row, value=A1)
        sheet.cell(column=8, row=sheet.max_row, value=T2)

        file.save('Customer_data.xlsx')

        messagebox.showinfo("info", "Sucessfully data entered!!!")

        Clear()

        registration_no()


# gender
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"

    else:
        gender = "Female"


# top frames
Label(root, text='', width=10, height=1, bg="#429683", font="arial 28 bold").pack(side=TOP, fill=X)
Label(root, text='AD Store', width=10, height=1, bg="#429683", font="arial 30 bold").pack(side=TOP, fill=X)
Label(root, text='!..Welcome..!', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
    side=TOP, fill=X)

# bottom frames
Label(root, text='', width=10, height=1, bg="#429683", ).pack(side=BOTTOM, fill=X)
Label(root, text='077-0867965', width=10, height=1, bg="#429683", font='arial 10 bold' ).pack(side=BOTTOM, fill=X)
Label(root, text='Galle', width=10, height=1, bg="#429683", font='arial 10 bold' ).pack(side=BOTTOM, fill=X)
Label(root, text='Karagoda', width=10, height=1, bg="#429683", font='arial 10 bold' ).pack(side=BOTTOM, fill=X)
Label(root, text='Mulanagoda,', width=10, height=1, bg="#429683", font='arial 10 bold' ).pack(side=BOTTOM, fill=X)

Label(root, text='Thank You Come Again', width=10, height=2, bg="#70BBAA", fg='#fff', font='arial 20 bold').pack(
    side=BOTTOM, fill=X)

# Registration and Date
Label(root, text="Registration No.", font='arial 13', fg=framebg, bg=background).place(x=30, y=200)
Label(root, text="Date", font='arial 13', fg=framebg, bg=background).place(x=500, y=200)
Label(root, text="Time", font='arial 13', fg=framebg, bg=background).place(x=700, y=200)

Registration = IntVar()
Date = StringVar()
Time = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font='arial 10')
reg_entry.place(x=160, y=200)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font='arial 10')
date_entry.place(x=550, y=200)

Date.set(d1)

#day = time.day()
#t1 = day.strftime("%H:%M:%S %p")
#time_entry = Entry(root, textvariable="Time", width=15, font='arial 10')
#time_entry.place(x=750, y=200)

#Time.set(t1)
# Customer details
obj = LabelFrame(root, text="Customer's Details", font=20, bd=3, width=900, bg=framebg, fg=framefg, height=250,
                 relief=GROOVE)
obj.place(x=50, y=250)

Label(root, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=90, y=320)
Label(root, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=90, y=370)
Label(root, text="NIC:", font="arial 13", bg=framebg, fg=framefg).place(x=90, y=420)

Label(root, text="Address:", font="arial 13", bg=framebg, fg=framefg).place(x=550, y=320)
Label(root, text="Time:", font="arial 13", bg=framebg, fg=framefg).place(x=550, y=370)
Label(root, text="Tel. No.:", font="arial 13", bg=framebg, fg=framefg).place(x=550, y=420)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=30, font="arial 10")
name_entry.place(x=180, y=50)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=180, y=100)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=240, y=100)

NIC = StringVar()
nic_entry = Entry(obj, textvariable=NIC, width=20, font="arial 10")
nic_entry.place(x=180, y=150)

# Time = StringVar()
# Label(obj, text="Time:", font="arial 13", bg=framebg, fg=framefg).place(x=630, y=50)
time_entry = Entry(obj, textvariable=Time, width=20, font="arial 10")
time_entry.place(x=630, y=100)

# current = time.current()
# t1 = current.strftime("%d/%m/%Y")
# time_entry = Entry(obj, textvariable=Time, width=15, font='arial 10')
# time_entry.place(x=550, y=200)


# Time.set(t1)
# te=Entry(obj,textvariable=Time, width=20, font="arial 10")
# te.place(x=630, y=50)

# def update():
#  t = strftime("%H:%M:%S %p")
# te.config(text=t)
# obj.after(1000, update)

# update()

# te.place(x=630, y=50)


Address = StringVar()
address_entry = Entry(obj, textvariable=Address, width=30,  font="arial 10")
address_entry.place(x=630, y=50)

Tel = StringVar()
tel_entry = Entry(obj, textvariable=Tel, width=20, font="arial 10")
tel_entry.place(x=630, y=150)

# button
Button(root, text="Sign in", width=19, height=2, font="arial 12 bold", bg="#81CBEB", command=ok).place(x=1000, y=200)

saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="#9BBEEA", command=Save)
saveButton.place(x=1000, y=280)

Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="#93E0EA", command=Clear).place(x=1000, y=360)

Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="#93CAEA", command=Exit).place(x=1000, y=440)

root.mainloop()
